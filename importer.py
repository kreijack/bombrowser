"""
BOM Browser - tool to browse a bom
Copyright (C) 2021 Goffredo Baroncelli <kreijack@inwind.it>

This program is free software; you can redistribute it and/or
modify it under the terms of the GNU General Public License
as published by the Free Software Foundation; either version 2
of the License, or (at your option) any later version.

This program is distributed in the hope that it will be useful,
but WITHOUT ANY WARRANTY; without even the implied warranty of
MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
GNU General Public License for more details.

You should have received a copy of the GNU General Public License
along with this program; if not, write to the Free Software
Foundation, Inc., 51 Franklin Street, Fifth Floor, Boston, MA  02110-1301, USA.
"""

import traceback
import sys
import pprint
import csv
import openpyxl
import xlrd

from PySide2.QtWidgets import QMessageBox, QFileDialog

import cfg
import db
import utils

_default_unit = "NR"

def import_csv_parent_child(keyword_map, options):

    (fn, _) = QFileDialog.getOpenFileName(None, "Select a file to import",
        None,
        "CSV files (*.csv);;Excel file (*.xls *.xlsx);;All files (*.*)")
    if fn == "":
        return None

    data = _read_table(fn)

    bom = _import_csv_parent_child2(data, keyword_map, options)
    return bom

def _read_table_xls(nf):
    book = xlrd.open_workbook(nf)
    sh = book.sheet_by_index(0)
    ret = []
    for row in range(sh.nrows):
        ret.append([sh.cell_value(row, col) for col in range(sh.ncols)])

    return ret

def _read_table_xlsx(nf):
    book = openpyxl.load_workbook(nf)
    sh = book.active
    ret = []
    for row in range(sh.max_row-1):
        ret.append([str(sh.cell(row+1, col+1).value)
            for col in range(sh.max_column)])

    return ret

def _read_table_csv(nf):

    with open(nf, newline='') as csvfile:
        c = csvfile.read(1)
        if c == '\ufeff':
            enc = 'utf-8-sig'
        else:
            enc = 'utf-8'

    if enc is None:
        csvfile = open(nf, newline='')
    else:
        csvfile = open(nf, newline='', encoding=enc)

    dialect = csv.Sniffer().sniff(csvfile.read(1024*16))
    csvfile.seek(0)
    reader = csv.reader(csvfile, dialect)
    reader = list(reader)

    if reader[0][0][0] == '\ufeff':
        reader[0][0] = reader[0][0][1:]

    return reader

def _read_table(nf):
    if nf.lower().endswith(".csv"):
        return _read_table_csv(nf)
    elif nf.lower().endswith(".xlsx"):
        return _read_table_xlsx(nf)
    elif nf.lower().endswith(".xls"):
        return _read_table_xls(nf)
    else:
        raise Exception("Format of file '%s' unknown; supperted file: .csv or .xls[x]"%(
            nf))

def _import_csv_parent_child2(data, keyword_map, options):

    separator=";"
    default_unit = "NR"
    ignore_duplicate = False
    ignore_quote = False
    ignore_bom = True
    skip_first_lines = 0

    if "separator" in options:
        separator=options["separator"]
    if "default_unit" in options:
        default_unit = options["default_unit"]
    if "ignore_duplicate" in options:
        ignore_duplicate = int(options["ignore_duplicate"]) > 0
    if "ignore_quote" in options:
        ignore_quote = int(options["ignore_quote"]) > 0
    if "ignore_bom" in options:
        ignore_bom = int(options["ignore_bom"]) > 0
    if "skip_first_lines" in options:
        skip_first_lines = int(options["skip_first_lines"])



    if separator=="COMMA":
        separator=","
    elif separator=="SEMICOLON":
        separator=";"
    elif separator=="COLON":
        separator=":"
    elif separator=="TAB":
        separator="\t"

    headers = data[skip_first_lines]

    if len(headers) < 1:
        raise Exception("Too few column")

    map1 = dict()
    colmap = {
        "code": None,
        "descr": None,
        "parent_code": None,
        "parent_descr": None,
        "qty" : None,
        "unit" : None,
        "each" : None,
        "ref": None,
        "ver": None,
    }
    for i in range(0, db.gvals_count):
        colmap["gval%d"%(i)] = None

    for v in keyword_map:
        s1, s2 = v.split("=")
        map1[s2] = s1

    for i, name in enumerate(headers):
        if name in map1:
            name = map1[name]
        if  name in colmap:
            colmap[name] = i

        # ignore other names

    if colmap["code"] is None or colmap["qty"] is None:
        raise Exception("Some mandatory columns are missing")

    linenr = skip_first_lines
    bom = dict()

    for fields in data[skip_first_lines+1:]:
        linenr += 1

        if len(fields) != len(headers):
            raise Exception("Error at line %d: the number of column is different from the header one"%(linenr))

        code_values = {
            "parent_code": 0,
            "unit": default_unit,
            "each": 1,
            "ref": ""
        }

        def xfloat(x):
            if isinstance(x, str):
                x = str(x).replace(",", ".")
            return float(x)



        for v in colmap:
            if colmap[v] is None:
                continue
            if v in ["qty", "each"]:
                code_values[v] = xfloat(fields[colmap[v]])
            else:
                code_values[v] = fields[colmap[v]]

        if not code_values["parent_code"] in bom:
            bom[code_values["parent_code"]] = {
                "deps": dict(),
                "code:": code_values["parent_code"],
            }
            if "parent_descr" in code_values:
                bom[code_values["parent_code"]]["descr"] = code_values["parent_descr"]

        if code_values["code"] in bom[code_values["parent_code"]]["deps"]:
            if not ignore_duplicate:
                raise Exception("Error at line %d: this line is a duplicated"%(linenr))

        bom[code_values["parent_code"]]["deps"][code_values["code"]] = {
            "code": code_values["code"],
            "qty": code_values["qty"],
            "each": code_values["each"],
            "unit": code_values["unit"],
            "ref": code_values["ref"],
        }

        if not code_values["code"] in bom:
            bom[code_values["code"]] = {"deps": dict()}

        for v in code_values:
            if v in ["parent_code", "parent_descr"]:
                continue
            if colmap[v] is None:
                continue
            bom[code_values["code"]][v] = code_values[v]



    return bom

def import_json(keyword_map):
    return { '0': {} }

def get_importer_list():

    l = cfg.config()["BOMBROWSER"].get("importer_list", "").split(",")
    if l == [""]:
        return None
    ret = []
    for importer_name in l:
        if not cfg.config().has_section(importer_name):
            continue
        name = cfg.config()[importer_name].get("name", None)
        type_ = cfg.config()[importer_name].get("type", None)
        map_ = cfg.config()[importer_name].get("map", None)

        if map_ is None or type_ is None or name is None:
            print("Importer '%s' is not defined properly"%(importer_name))
            continue

        if not type_ in ["parent_child", "json"]:
            print("Importer '%s': unknown type '%s'"%(importer_name, type_))
            continue

        name, map_, type_ = map(lambda x : x.strip(), (name, map_, type_))

        if type_ == "parent_child":
            separator = cfg.config()[importer_name].get("separator", ";")
            optionss = cfg.config()[importer_name].get("options", "")
            optionsl = optionss.split(",")
            options = {
                "separator": separator,
            }
            if optionss != "":
                for (k,v) in [x.split("=") for x in optionsl]:
                    options[k]=v
            ret.append((importer_name, name,
                utils.Callable(import_csv_parent_child, map_.split(","), options)))
        else: # json
            ret.append((importer_name, name,
                utils.Callable(import_json, map_)))

    return ret

def test_import_csv_parent_child2():
    import io
    data = """parent_code;parent_descr;code;descr;qty;each;unit;gval1
Z;0-descr;1;1-descr;1;2;nr;gval-1
Z;0-descr;2;2-descr;2;3;nr;gval-2
4;4-descr;3;3-descr;8;9;nr;gval-7
Z;0-descr;4;4-descr;1.5;3;gr;gval-4"""

    data = [x.split(";") for x in data.split("\n")]

    bom = _import_csv_parent_child2(data,[], {})

    assert("Z" in bom)
    assert("1" in bom)

    assert(bom["Z"]["descr"] == "0-descr")
    assert(bom["1"]["descr"] == "1-descr")
    assert(bom["1"]["gval1"] == "gval-1")

    assert("1" in bom["Z"]["deps"])
    assert(bom["Z"]["deps"]["1"]["qty"] == 1)
    assert(bom["Z"]["deps"]["1"]["each"] == 2)
    assert(bom["Z"]["deps"]["1"]["unit"] == "nr")
    assert(bom["Z"]["deps"]["4"]["qty"] == 1.5)

    assert("2" in bom["Z"]["deps"])
    assert("4" in bom["Z"]["deps"])

    assert("3" in bom["4"]["deps"])

    assert("4" in bom["Z"]["deps"])

def test_import_csv_parent_child2_error_on_duplicate_row():
    import io
    s = """parent_code;parent_descr;code;descr;qty;each;unit;gval1
0;0-descr;1;1-descr;1;2;nr;gval-1
0;0-descr;2;2-descr;2;3;nr;gval-2
4;4-descr;3;3-descr;8;9;nr;gval-7
0;0-descr;4;4-descr;1.5;3;gr;gval-4"""

    data = [x.split(";") for x in s.split("\n")]
    bom = _import_csv_parent_child2(data,[], {})

    # ok the file is ok

    # add a duplicate
    s1 = s + '\n0;0-descr;4;4-descr;1.5;3;gr;gval-4'
    data = [x.split(";") for x in s1.split("\n")]
    failed = True
    try:
        f = io.StringIO(s1)
        bom = _import_csv_parent_child2(data,[], {})
    except:
        failed = False

    assert(not failed)

    s1 = s + '\n0;0-descr;4;4-descr;1.5;3;gr;gval-4'
    data = [x.split(";") for x in s1.split("\n")]
    bom = _import_csv_parent_child2(data,[], {"ignore_duplicate":"1"})

def test_import_csv_parent_child2_error_on_less_columns():
    import io
    s = """parent_code;parent_descr;code;descr;qty;each;unit;gval1
0;0-descr;1;1-descr;1;2;nr;gval-1
0;0-descr;2;2-descr;2;3;nr;gval-2
4;4-descr;3;3-descr;8;9;nr;gval-7
0;0-descr;4;4-descr;1.5;3;gr;gval-4"""

    data = [x.split(";") for x in s.split("\n")]
    bom = _import_csv_parent_child2(data,[], {})

    # ok the file is ok

    # add a row with less columns
    s1 = s + '\n0;0-descr;7;4-descr;1.5;3;gr'
    failed = True
    try:
        data = [x.split(";") for x in s1.split("\n")]
        bom = _import_csv_parent_child2(data, [], {})
    except:
        failed = False

    assert(not failed)

    s2 = s1 + ';"gval-4"'
    data = [x.split(";") for x in s2.split("\n")]
    bom = _import_csv_parent_child2(data, [], {})

def test_import_csv_parent_child2_error_on_more_columns():
    import io
    s = """parent_code;parent_descr;code;descr;qty;each;unit;gval1
0;0-descr;1;1-descr;1;2;nr;gval-1
0;0-descr;2;2-descr;2;3;nr;gval-2
4;4-descr;3;3-descr;8;9;nr;gval-7
0;0-descr;4;4-descr;1.5;3;gr;gval-4"""

    data = [x.split(";") for x in s.split("\n")]
    bom = _import_csv_parent_child2(data,[], {})

    # ok the file is ok

    # add a row with the same column columns
    s1 = s + '\n0;0-descr;7;4-descr;1.5;3;gr;gval-4'

    data = [x.split(";") for x in s1.split("\n")]
    bom = _import_csv_parent_child2(data, [], {})

    s2 = s1 + ';gval-4'
    failed = True
    try:
        data = [x.split(";") for x in s2.split("\n")]
        bom = _import_csv_parent_child2(data,[], {})
    except:
        failed = False

    assert(not failed)

def test_import_csv_parent_child2_wo_parent_columns():
    import io
    s = """code\tdescr\tqty\teach\tunit\tgval1
1\t1-descr\t1\t2\tnr\tgval-1
2\t2-descr\t2\t3\tnr\tgval-2
3\t3-descr\t8\t9\tnr\tgval-7
4\t4-descr\t1.5\t3\tgr\tgval-4"""

    data = [x.split("\t") for x in s.split("\n")]
    bom = _import_csv_parent_child2(data,[], {})

    assert(0 in bom)
    assert("1" in bom)

    assert( bom[0]["deps"]["2"]["qty"] == 2)
    assert( bom[0]["deps"]["2"]["each"] == 3)


if __name__ == "__main__":
    last = 1
    import test_db
    test_db.run_test(sys.argv[last:], sys.modules[__name__])
